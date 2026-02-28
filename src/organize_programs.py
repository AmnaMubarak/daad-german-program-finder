import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'output')

input_file = os.path.join(OUTPUT_DIR, 'german_cs_masters_programs.xlsx')
df = pd.read_excel(input_file)

print(f"Total programs: {len(df)}")
print(f"\nColumns: {df.columns.tolist()}")

def categorize_program(program_name):
    if pd.isna(program_name):
        return "Uncategorized"
    
    program_name = str(program_name).lower()
    
    if any(keyword in program_name for keyword in ['data science', 'data analytics', 'artificial intelligence', 'machine learning', 'big data', 'data engineering']):
        return "Data Science & AI"
    elif any(keyword in program_name for keyword in ['cyber', 'security', 'information security']):
        return "Cybersecurity"
    elif any(keyword in program_name for keyword in ['software engineering', 'software development', 'software systems']):
        return "Software Engineering"
    elif any(keyword in program_name for keyword in ['business', 'management', 'entrepreneurship', 'innovation', 'digital transformation', 'information systems']):
        return "Business & IT"
    elif any(keyword in program_name for keyword in ['web', 'mobile', 'internet', 'digital media']):
        return "Web & Mobile Development"
    elif any(keyword in program_name for keyword in ['network', 'communication', 'telecommunication', 'distributed']):
        return "Networks & Communications"
    elif any(keyword in program_name for keyword in ['robot', 'embedded', 'automation', 'mechatronic']):
        return "Robotics & Embedded Systems"
    elif any(keyword in program_name for keyword in ['game', 'gaming']):
        return "Game Development"
    elif any(keyword in program_name for keyword in ['human-computer', 'interaction', 'human computer', 'media informatics']):
        return "HCI & UX"
    elif any(keyword in program_name for keyword in ['bioinformatics', 'computational biology', 'medical informatics']):
        return "Bioinformatics"
    elif any(keyword in program_name for keyword in ['computer science', 'computing', 'informatics', 'information technology']):
        return "Computer Science (General)"
    elif any(keyword in program_name for keyword in ['computational', 'scientific computing']):
        return "Computational Science"
    else:
        return "Other Tech Programs"

df['Category'] = df['Program Name'].apply(categorize_program)

print("\n" + "="*60)
print("PROGRAM CATEGORIES DISTRIBUTION")
print("="*60)
category_counts = df['Category'].value_counts()
for category, count in category_counts.items():
    print(f"{category}: {count} programs")

category_colors = {
    "Data Science & AI": "B4C7E7",
    "Computer Science (General)": "C5E0B4",
    "Cybersecurity": "F8CBAD",
    "Software Engineering": "FFE699",
    "Business & IT": "E2EFDA",
    "Web & Mobile Development": "FCE4D6",
    "Networks & Communications": "D9E1F2",
    "Robotics & Embedded Systems": "EDEDED",
    "Game Development": "F4B084",
    "HCI & UX": "C9DAF8",
    "Bioinformatics": "D5A6BD",
    "Computational Science": "B6D7A8",
    "Other Tech Programs": "FFFFFF",
    "Uncategorized": "F2F2F2"
}

output_file = os.path.join(OUTPUT_DIR, 'german_cs_masters_programs_organized.xlsx')

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_sorted = df.sort_values(['Category', 'Program Name'])
    df_sorted.to_excel(writer, sheet_name='All Programs', index=False)
    
    for category in sorted(df['Category'].unique()):
        category_df = df[df['Category'] == category].copy()
        category_df_display = category_df.drop('Category', axis=1)
        sheet_name = category[:31] if len(category) <= 31 else category[:28] + "..."
        category_df_display.to_excel(writer, sheet_name=sheet_name, index=False)

wb = load_workbook(output_file)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

ws_all = wb['All Programs']

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for cell in ws_all[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

for row_idx, row in enumerate(ws_all.iter_rows(min_row=2), start=2):
    category_col = None
    for col_idx, cell in enumerate(ws_all[1], start=1):
        if cell.value == 'Category':
            category_col = col_idx
            break
    
    if category_col:
        category_value = ws_all.cell(row=row_idx, column=category_col).value
        if category_value in category_colors:
            row_fill = PatternFill(start_color=category_colors[category_value], 
                                   end_color=category_colors[category_value], 
                                   fill_type="solid")
            for cell in row:
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

column_widths = {
    'A': 45, 'B': 35, 'C': 20, 'D': 15, 'E': 18,
    'F': 15, 'G': 22, 'H': 12, 'I': 15, 'J': 15,
    'K': 30, 'L': 50, 'M': 20, 'N': 25
}

for col, width in column_widths.items():
    ws_all.column_dimensions[col].width = width

ws_all.freeze_panes = 'A2'

for sheet_name in wb.sheetnames:
    if sheet_name != 'All Programs':
        ws = wb[sheet_name]
        
        category_color = "FFFFFF"
        for cat, color in category_colors.items():
            if cat.startswith(sheet_name.replace("...", "")):
                category_color = color
                break
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row_fill = PatternFill(start_color=category_color, 
                               end_color=category_color, 
                               fill_type="solid")
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        for col, width in column_widths.items():
            if col != 'N':
                ws.column_dimensions[col].width = width
        
        ws.freeze_panes = 'A2'

wb.save(output_file)

print(f"\n{'='*60}")
print(f"✓ Successfully created: {output_file}")
print(f"{'='*60}")
print(f"\nFile contains {len(wb.sheetnames)} sheets:")
for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {idx}. {sheet_name}")

print("\n✓ Color-coded categories applied")
print("✓ Each category has its own sheet")
print("✓ Headers formatted with blue background")
print("✓ Columns auto-sized for readability")
print("✓ Borders applied to all cells")
print("✓ Freeze panes enabled for easy scrolling")
